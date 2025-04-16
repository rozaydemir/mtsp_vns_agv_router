from genetic import Genetic
from heuristic import Heuristic
from algorithm import Algorithm
from rule_of_thumb_api import RuleOfThumb
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import time


class ScenarioAnalysis:
    def __init__(self):
        self.INSTANCES = [
            # "Instances/vestel-fabrika-verisi.txt",
            "Instances/2xVestel-synomim.txt",
            # "Instances/lrc5.txt",
            # "Instances/lrc11.txt",
            # "Instances/lrc11-location-based.txt",
            # "Instances/lrc11-demand-based.txt",
            # "Instances/lrc11-time-window-based.txt",
            # "Instances/lrc11-servtime-based.txt",
            # "Instances/lrc15.txt",
            # "Instances/lrc15-location-based.txt",
            # "Instances/lrc15-demand-based.txt",
            # "Instances/lrc15-time-window-based.txt",
            # "Instances/lrc15-servtime-based.txt",
            # "Instances/lrc19.txt",
            # "Instances/lrc19-location-based.txt",
            # "Instances/lrc19-time-window-based.txt",
            # "Instances/lrc19-servtime-based.txt",
            # "Instances/lrc19-demand-based.txt"
        ]
        self.VEHICLE_COUNT = [
            10
            # ,
            # 2
            # ,
            # 3
        ]
        self.TROLLEY_COUNT = [
            # 1,
            # 2,
            3
        ]
        self.TROLLEY_IMPACT_RATE = [
            # 20,
            30,
            # 40
        ]
        self.EARLINESS_TARDINESS_PENALTY = [
            # 10,
            20,
            # 30
        ]

    def parse_route_math_model(self, solutionResult_MATH_MODEL, depotNode):
        resultList = list()
        for i in range(0, len(solutionResult_MATH_MODEL)):
            worksheetMathModelListItem = list()
            vehicle_info = f"Vehicle {i} - Trolley Count: {solutionResult_MATH_MODEL[i]['trolleyCount']}"
            route_math_model = ' - '.join(solutionResult_MATH_MODEL[i]['route'])

            pairs = route_math_model.split(" - ")
            pairs = [eval(pair) for pair in pairs]
            route_map = {}

            for start, end in pairs:
                route_map[start] = end

            start_point = pairs[0][0]
            for start, end in pairs:
                if start not in route_map.values():
                    start_point = start
                    break

            route = ["D0"]
            current = start_point
            while current in route_map:
                next_point = route_map[current]
                if next_point == depotNode:
                    route.append("D0")
                else:
                    route.append(f"C{next_point}")
                current = next_point

            route_str = " - ".join(map(str, route))
            worksheetMathModelListItem.append(vehicle_info)
            worksheetMathModelListItem.append(route_str)
            resultList.append(" => ".join(worksheetMathModelListItem))

        return " ; ".join(resultList)

    def cost_detail_math_model(self, routesCostDetailResult):
        distCost = sum(item[0] for item in routesCostDetailResult)
        EAPenalty = sum(item[1] for item in routesCostDetailResult if item[1] >= 0)
        TAPenalty = sum(item[2] for item in routesCostDetailResult if item[2] >= 0)
        return distCost, EAPenalty, TAPenalty

    def cost_detail_heuristic(self, solutionResult_ALNS):
        distCost = 0
        EAPenalty = 0
        TAPenalty = 0
        for k in range(0, len(solutionResult_ALNS)):
            if len(solutionResult_ALNS[k]) > 0:
                distCost += sum(item[0] for item in solutionResult_ALNS[k]["costDetail"])
                EAPenalty += sum(item[1] for item in solutionResult_ALNS[k]["costDetail"] if item[1] > 0)
                TAPenalty += sum(item[2] for item in solutionResult_ALNS[k]["costDetail"] if item[2] > 0)

        return distCost, EAPenalty, TAPenalty

    def parse_route_alns(self, solutionResult_ALNS):
        resultList = list()
        for k in range(0, len(solutionResult_ALNS)):
            if len(solutionResult_ALNS[k]) > 0:
                worksheetALNSListItem = list()
                vehicle_info_alns = f"Vehicle {k} - Trolley Count: {solutionResult_ALNS[k]['trolleyCount']}"
                route_alns = ' - '.join(solutionResult_ALNS[k]['route'])
                worksheetALNSListItem.append(vehicle_info_alns)
                worksheetALNSListItem.append(route_alns)
                resultList.append(" => ".join(worksheetALNSListItem))
        return " ; ".join(resultList)

    def parse_detail_excel(self, worksheetDetail, iterationNumber, trolleyCount, trolleyImpactRate,
                           earlinessTardinessPenalty,
                           bestCost_MATH_MODEL, cpuTime_MATH_MODEL, routesDetailResult,
                           bestCost_ALNS, cpuTime_ALNS, solutionResult_ALNS,
                           bestCost_GA, cpuTime_GA, solutionResult_GA,
                           instance,
                           cpuTime_RULE_OF_THUMB, bestCost_RULE_OF_THUMB, solutionResult_RULE_OF_THUMB,
                           gapAlnsToMilp, gapAlnsToGA
                           ):
        worksheetDetail.append(
            ["ID", 'TROLLEY_COUNT', 'TROLLEY_IMPACT_RATE', 'EARLINESS/TARDINESS PENALTY'])
        worksheetDetail.append(["TEST_ID_" + str(iterationNumber), trolleyCount, trolleyImpactRate,
                                earlinessTardinessPenalty])
        worksheetDetail.append(
            [instance, "Mathematical Formulation", "ALNS", "GA", "RULE OF THUMB", "GAP (ALNS-MILP)", "GAP (ALNS-GA)"])
        worksheetDetail.append(
            ["Result", round(bestCost_MATH_MODEL), round(bestCost_ALNS), round(bestCost_GA), bestCost_RULE_OF_THUMB,
             gapAlnsToMilp, gapAlnsToGA])
        worksheetDetail.append(["cpuTime", cpuTime_MATH_MODEL, cpuTime_ALNS, cpuTime_GA, cpuTime_RULE_OF_THUMB])
        worksheetDetail.append(["Mathematical Formulation"])
        for rd in range(0, len(routesDetailResult)):
            worksheetDetail.append(["", routesDetailResult[rd]])

        worksheetDetail.append(["ALNS Algorithm"])
        for k in range(0, len(solutionResult_ALNS)):
            if len(solutionResult_ALNS[k]) > 0:
                vehicle_info_alns = f"Vehicle {k} - Trolley Count: {solutionResult_ALNS[k]['trolleyCount']}"
                worksheetDetail.append(["", vehicle_info_alns])
                for rda in range(0, len(solutionResult_ALNS[k]["routeDetail"])):
                    worksheetDetail.append(["", solutionResult_ALNS[k]["routeDetail"][rda]])

        worksheetDetail.append(["GA Algorithm"])
        for k in range(0, len(solutionResult_GA)):
            if len(solutionResult_GA[k]) > 0:
                vehicle_info_alns = f"Vehicle {k} - Trolley Count: {solutionResult_GA[k]['trolleyCount']}"
                worksheetDetail.append(["", vehicle_info_alns])
                for rda in range(0, len(solutionResult_GA[k]["routeDetail"])):
                    worksheetDetail.append(["", solutionResult_GA[k]["routeDetail"][rda]])

        worksheetDetail.append(["Rule Of Thumb Algorithm"])
        for kr in range(0, len(solutionResult_RULE_OF_THUMB)):
            if len(solutionResult_RULE_OF_THUMB[kr]) > 0:
                vehicle_info_alns = f"Vehicle {kr} - Trolley Count: {solutionResult_RULE_OF_THUMB[kr]['trolleyCount']}"
                worksheetDetail.append(["", vehicle_info_alns])
                for rdaT in range(0, len(solutionResult_RULE_OF_THUMB[kr]["routeDetail"])):
                    worksheetDetail.append(["", solutionResult_RULE_OF_THUMB[kr]["routeDetail"][rdaT]])

        worksheetDetail.append([])

    def parse_roth_detail_excel(self, worksheetRot, file_name, vehicleCount, iterationNumber, trolleyCount,
                                trolleyImpactRate, earlinessTardinessPenalty,
                                minResult, rule_of_thumb_dist_cost, rule_of_thumb_EACost, rule_of_thumb_TAcost,
                                rule_of_thumb_routes, bestCost_RULE_OF_THUMB, bestDistance_TWO,
                                bestDistance_THREE, bestDistance_FOUR, milpresult, alnsresult
                                ):
        worksheetRot.append(
            [
                iterationNumber, file_name, vehicleCount, trolleyCount, trolleyImpactRate,
                earlinessTardinessPenalty,
                minResult[2], round(minResult[1]),
                round(rule_of_thumb_dist_cost),
                round(rule_of_thumb_EACost), round(rule_of_thumb_TAcost), rule_of_thumb_routes,
                minResult[0],
                round(bestCost_RULE_OF_THUMB),
                round(bestDistance_TWO),
                round(bestDistance_THREE),
                round(bestDistance_FOUR),
                0,
                0
            ])

    def execute(self):
        start_time = time.time()
        print("Scenario Started")
        iterationNumber = 0
        iterationCount = 12000
        capacityOfTrolley = 60
        timeLimit = 60000 * 20
        file_name = "result-all-ga-with-new-params-4"
        workbook = Workbook()
        workbook.remove(workbook.active)  # Remove default sheet
        worksheet = workbook.create_sheet(title=f'Test Result')
        worksheet.append(
            [
                "", "TEST INSTANCES", "", "", "", "",
                # "MATHEMATICAL FORMULATION", "", "", "", "", "", "",
                "ALNS", "", "", "", "", "",
                "Genetic Algorithm", "", "", "", "", "",
                # "Min Rule Of Thumb", "", "", "", "", "", "",
                # "GAP (ALNS-MILP)",
                # "GAP (ALNS-GA)",
            ])
        worksheet.append(
            [
                "TEST ID", "FILE NAME", "VEHICLE COUNT", 'TROLLEY COUNT', 'TROLLEY IMPACT TIME',
                "EARLINESS/TARDINESS PENALTY",
                # "Math Form Is Optimal OR Feasible", "CPU TIME", "RESULT", "Math Model Distance Cost",
                # "Math Model EA Cost", "Math Model TA Cost", "ROUTES",
                # "CPU TIME", "RESULT",
                # "ALNS Distance Cost", "ALNS EA Cost", "ALNS TA Cost", "ROUTES",
                # "CPU TIME", "RESULT",
                "GA Distance Cost", "GA EA Cost", "GA TA Cost", "ROUTES",
                "CPU TIME", "RESULT",
                # "ROTH Distance Cost", "ROTH EA Cost", "ROTH TA Cost", "ROUTES", "MIN ALGO NAME",
                # "",
                # ""
            ])
        infeasibleData = list()
        nonOptimalSolution = list()
        workbookDetail = Workbook()
        workbookDetail.remove(workbookDetail.active)  # Remove default sheet

        workbookROTDetail = Workbook()
        workbookROTDetail.remove(workbookROTDetail.active)  # Remove default sheet
        worksheetRot = workbookROTDetail.create_sheet(title=f'Test Result')
        # worksheetRot.append(
        #     [
        #         "", "TEST INSTANCES", "", "", "", "",
        #         "Min Rule Of Thumb", "", "", "", "", "",
        #         "", "", "", "", "", "", ""
        #     ])
        # worksheetRot.append(
        #     [
        #         "TEST ID", "FILE NAME", "VEHICLE COUNT", 'TROLLEY COUNT', 'TROLLEY IMPACT TIME',
        #         "EARLINESS/TARDINESS PENALTY",
        #         "CPU TIME", "RESULT",
        #         "ROTH Distance Cost", "ROTH EA Cost", "ROTH TA Cost", "ROUTES", "MIN ALGO NAME",
        #         "SORTED BY COST", "GREEDY", "SORTED BY DIST", "SORTED BY DUE DATE", "MILP Result", "ALNS result"
        #     ])


        for ins in range(0, len(self.INSTANCES)):
            instance = self.INSTANCES[ins]
            fileName = instance.split('/')[-1].replace('.txt', '')
            for vc in range(0, len(self.VEHICLE_COUNT)):
                vehicleCount = self.VEHICLE_COUNT[vc]
                worksheetDetail = workbookDetail.create_sheet(title=f'{vehicleCount} Vehicle')
                for tc in range(0, len(self.TROLLEY_COUNT)):
                    trolleyCount = self.TROLLEY_COUNT[tc]
                    for tir in range(0, len(self.TROLLEY_IMPACT_RATE)):
                        trolleyImpactRate = self.TROLLEY_IMPACT_RATE[tir]
                        for etp in range(0, len(self.EARLINESS_TARDINESS_PENALTY)):
                            iterationNumber += 1
                            earlinessTardinessPenalty = self.EARLINESS_TARDINESS_PENALTY[etp]
                            print(
                                "===================================================================================================================================")
                            print("Iteration Number : " + str(iterationNumber))
                            print(
                                f"INSTANCE: {instance}, VEHICLE_COUNT : {vehicleCount}, TROLLEY_COUNT: {trolleyCount}"
                                f", TROLLEY_IMPACT_RATE: {trolleyImpactRate}, EARLINESS_TARDINESS_PENALTY: {earlinessTardinessPenalty}, "
                                f"POPULATION SIZE: {40}, SELECTION SIZE: {8}")

                            # print("==================== MILP START =====================")

                            # algorithm = Algorithm(instance, vehicleCount, capacityOfTrolley, trolleyCount,
                            #                       trolleyImpactRate, earlinessTardinessPenalty,
                            #                       earlinessTardinessPenalty, timeLimit)
                            # optimalOrFeasible, bestCost_MATH_MODEL, cpuTime_MATH_MODEL, solutionResult_MATH_MODEL, depotNode, routesDetailResult, routesCostDetailResult = algorithm.execute()
                            #
                            # if optimalOrFeasible == "NON-OPTIMAL":
                            #     nonOptimalSolution.append(
                            #         ["TEST_ID: " + str(iterationNumber),
                            #          "INSTANCE: " + str(
                            #              instance) + ", VEHICLE_COUNT: " + str(
                            #              vehicleCount) + ", TROLLEY_COUNT: " + str(
                            #              trolleyCount) + ", TROLLEY_IMPACT_RATE: " + str(
                            #              trolleyImpactRate) + ", EARLINESS_TARDINESS_PENALTY: " + str(
                            #              earlinessTardinessPenalty)
                            #          ])

                            # print("==================== ALNS START =====================")
                            # heuristic = Heuristic(instance, vehicleCount, capacityOfTrolley, trolleyCount,
                            #                       trolleyImpactRate, earlinessTardinessPenalty,
                            #                       earlinessTardinessPenalty, iterationCount, 100, file_name)
                            # bestCost_ALNS, cpuTime_ALNS, solutionResult_ALNS = heuristic.execute()

                            print("==================== GA START =======================")
                            genetic = Genetic(instance, vehicleCount, capacityOfTrolley, trolleyCount,
                                              trolleyImpactRate, earlinessTardinessPenalty,
                                              earlinessTardinessPenalty, iterationCount, 40, 8)
                            bestCost_GA, cpuTime_GA, solutionResult_GA = genetic.execute()

                            # print("==================== ROTH START =====================")
                            # ruleOfThumb = RuleOfThumb(instance, vehicleCount, capacityOfTrolley, trolleyCount,
                            #                           trolleyImpactRate, earlinessTardinessPenalty,
                            #                           earlinessTardinessPenalty, 0, 100)
                            # bestCost_RULE_OF_THUMB, bestDistance_TWO, bestDistance_THREE, bestDistance_FOUR, minResult = ruleOfThumb.execute()

                            # gapAlnsToMilp = ((round(bestCost_MATH_MODEL) - round(bestCost_ALNS)) / round(
                            #     bestCost_ALNS)) * 100
                            # if gapAlnsToMilp > 1 or gapAlnsToMilp < 0:
                            #     infeasibleData.append(
                            #         ["TEST_ID: " + str(iterationNumber), "INSTANCE: " + str(
                            #             instance) + ", VEHICLE_COUNT: " + str(
                            #             vehicleCount) + ", TROLLEY_COUNT: " + str(
                            #             trolleyCount) + ", TROLLEY_IMPACT_RATE: " + str(
                            #             trolleyImpactRate) + ", EARLINESS_TARDINESS_PENALTY: " + str(
                            #             earlinessTardinessPenalty), gapAlnsToMilp, "ALNS TO MILP"])
                            #
                            # gapAlnsToGA = ((round(bestCost_GA) - round(bestCost_ALNS)) / round(
                            #     bestCost_ALNS)) * 100
                            # if gapAlnsToGA > 1 or gapAlnsToGA < 0:
                            #     infeasibleData.append(
                            #         ["TEST_ID: " + str(iterationNumber), "INSTANCE: " + str(
                            #             instance) + ", VEHICLE_COUNT: " + str(
                            #             vehicleCount) + ", TROLLEY_COUNT: " + str(
                            #             trolleyCount) + ", TROLLEY_IMPACT_RATE: " + str(
                            #             trolleyImpactRate) + ", EARLINESS_TARDINESS_PENALTY: " + str(
                            #             earlinessTardinessPenalty), gapAlnsToGA, "ALNS TO GA"])

                            # math_model_routes = self.parse_route_math_model(solutionResult_MATH_MODEL, depotNode)
                            # heuristic_routes = self.parse_route_alns(solutionResult_ALNS)
                            ga_routes = self.parse_route_alns(solutionResult_GA)
                            # rule_of_thumb_routes = self.parse_route_alns(minResult[3])

                            # mathModel_dist_cost, mathModel_EACost, mathModel_TAcost = self.cost_detail_math_model(
                            #     routesCostDetailResult)
                            # heuristic_dist_cost, heuristic_EACost, heuristic_TAcost = self.cost_detail_heuristic(
                            #     solutionResult_ALNS)
                            ga_dist_cost, ga_EACost, ga_TAcost = self.cost_detail_heuristic(
                                solutionResult_GA)
                            # rule_of_thumb_dist_cost, rule_of_thumb_EACost, rule_of_thumb_TAcost = self.cost_detail_heuristic(
                            #     minResult[3])
                            worksheet.append(
                                [
                                    iterationNumber, fileName, vehicleCount, trolleyCount, trolleyImpactRate,
                                    earlinessTardinessPenalty,
                                    # optimalOrFeasible,
                                    # cpuTime_MATH_MODEL, round(bestCost_MATH_MODEL), round(mathModel_dist_cost),
                                    # round(mathModel_EACost), round(mathModel_TAcost),
                                    # math_model_routes,
                                    # cpuTime_ALNS, round(bestCost_ALNS), round(heuristic_dist_cost),
                                    # round(heuristic_EACost), round(heuristic_TAcost), heuristic_routes,
                                    cpuTime_GA, round(bestCost_GA), round(ga_dist_cost),
                                    round(ga_EACost), round(ga_TAcost), ga_routes,
                                    # minResult[2], round(minResult[1]),
                                    # round(rule_of_thumb_dist_cost),
                                    # round(rule_of_thumb_EACost), round(rule_of_thumb_TAcost), rule_of_thumb_routes,
                                    # minResult[0],
                                    # gapAlnsToMilp,
                                    # gapAlnsToGA

                                ])
                            # self.parse_detail_excel(worksheetDetail, iterationNumber, trolleyCount,
                            #                         trolleyImpactRate, earlinessTardinessPenalty,
                            #                         bestCost_MATH_MODEL,  cpuTime_MATH_MODEL, routesDetailResult,
                            #                         bestCost_ALNS, cpuTime_ALNS, solutionResult_ALNS,
                            #                         bestCost_GA, cpuTime_GA, solutionResult_GA,
                            #                         instance,
                            #                         minResult[2], minResult[1],
                            #                         minResult[3],
                            #                         gapAlnsToMilp,
                            #                         gapAlnsToGA
                            #                         )
                            # self.parse_roth_detail_excel(worksheetRot, file_name, vehicleCount, iterationNumber,
                            #                              trolleyCount,
                            #                              trolleyImpactRate, earlinessTardinessPenalty,
                            #                              minResult,
                            #                              round(rule_of_thumb_dist_cost),
                            #                              round(rule_of_thumb_EACost), round(rule_of_thumb_TAcost),
                            #                              rule_of_thumb_routes,
                            #                              round(bestCost_RULE_OF_THUMB),
                            #                              round(bestDistance_TWO),
                            #                              round(bestDistance_THREE),
                            #                              round(bestDistance_FOUR), round(0),
                            #                              round(0))
            # workbookDetail.save(f'excels/details/{fileName}-details.xlsx')
            # workbookROTDetail.save(f'excels/rot-detail/{file_name}-roth-details.xlsx')

        # if len(infeasibleData) > 0:
        #     workbookInFeasible = Workbook()
        #     workbookInFeasible.remove(workbookInFeasible.active)
        #     worksheetInFeasible = workbookInFeasible.create_sheet(title="Gap Data")
        #     worksheetInFeasible.append(["TEST_ID", "TEST INSTANCE", "GAP"])
        #     for inf in range(0, len(infeasibleData)):
        #         worksheetInFeasible.append(infeasibleData[inf])
        #     workbookInFeasible.save(f'excels/gap/{file_name}-gap.xlsx')
        #
        #
        # if len(nonOptimalSolution) > 0:
        #     workbookNonOptimal = Workbook()
        #     workbookNonOptimal.remove(workbookNonOptimal.active)
        #     worksheetNonOptimal = workbookNonOptimal.create_sheet(title="Non-optimal Data")
        #     worksheetNonOptimal.append(["TEST_ID", "TEST INSTANCE"])
        #     for f in range(0, len(nonOptimalSolution)):
        #         worksheetNonOptimal.append(nonOptimalSolution[f])
        #     workbookNonOptimal.save(f'excels/non-optimal/{file_name}-non-optimal.xlsx')

        workbook.save(f'excels/{file_name}.xlsx')
        print("Scenario finished")
        endtime = time.time()
        cpuTime = round(endtime - start_time, 1)

        print("cpuTime: " + str(cpuTime) + " seconds")


scenario = ScenarioAnalysis()
scenario.execute()
