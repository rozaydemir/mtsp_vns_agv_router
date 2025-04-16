from genetic import Genetic
from heuristic import Heuristic
from algorithm import Algorithm
from rule_of_thumb_api import RuleOfThumb
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import time


class ScenarioAnalysis:
    def __init__(self):
        self.nonit = [163, 164, 165, 166, 167, 168, 169, 170, 171, 190, 191, 192, 193, 194, 195, 196, 197, 198, 217, 218, 219, 220, 221, 222, 223, 224, 225, 568, 569, 570, 571, 572, 573, 574, 575, 576, 588, 593, 595, 596, 597, 598, 599, 600, 601, 602, 603, 622, 623, 624, 625, 626, 627, 628, 629, 630, 636, 1159, 1160, 1161, 1162, 1163, 1164, 1165, 1166, 1167, 1168, 1169, 1170, 1171, 1172, 1173, 1174, 1175, 1176, 1177, 1178, 1179, 1182, 1183, 1186, 1189, 1190, 1191, 1192, 1193, 1194, 1195, 1196, 1197, 1198, 1199, 1200, 1201, 1202, 1203, 1204, 1205, 1206, 1210, 1213, 1214, 1215]
        self.initeration = [
    811, 812, 813, 814, 815, 816, 817, 818, 819, 820, 821, 822, 823, 824, 825, 826, 827, 828, 829, 830,
    831, 832, 833, 834, 835, 836, 837, 838, 839, 840, 841, 842, 843, 844, 845, 846, 847, 848, 849, 850,
    851, 852, 853, 854, 855, 856, 857, 858, 859, 860, 861, 862, 863, 864, 865, 866, 867, 868, 869, 870,
    871, 872, 873, 874, 875, 876, 877, 878, 879, 880, 881, 882, 883, 884, 885, 886, 887, 888, 889, 890,
    891, 892, 893, 894, 895, 896, 897, 898, 899, 900, 901, 902, 903, 904, 905, 906, 907, 908, 909, 910,
    911, 912, 913, 914, 915, 916, 917, 918, 919, 920, 921, 922, 923, 924, 925, 926, 927, 928, 929, 930,
    931, 932, 933, 934, 935, 936, 937, 938, 939, 940, 941, 942, 943, 944, 945, 946, 947, 948, 949, 950,
    951, 952, 953, 954, 955, 956, 957, 958, 959, 960, 961, 962, 963, 964, 965, 966, 967, 968, 969, 970,
    971, 972, 973, 974, 975, 976, 977, 978, 979, 980, 981, 982, 983, 984, 985, 986, 987, 988, 989, 990,
    991, 992, 993, 994, 995, 996, 997, 998, 999, 1000, 1001, 1002, 1003, 1004, 1005, 1006, 1007, 1008,
    1009, 1010, 1011, 1012, 1013, 1014, 1015, 1016, 1017, 1018, 1019, 1020, 1021, 1022, 1023, 1024,
    1025, 1026, 1027, 1028, 1029, 1030, 1031, 1032, 1033, 1034, 1035, 1036, 1037, 1038, 1039, 1040,
    1041, 1042, 1043, 1044, 1045, 1046, 1047, 1048, 1049, 1050, 1051, 1052, 1053, 1054, 1055, 1056,
    1057, 1058, 1059, 1060, 1061, 1062, 1063, 1064, 1065, 1066, 1067, 1068, 1069, 1070, 1071, 1072,
    1073, 1074, 1075, 1076, 1077, 1078, 1079, 1080, 1081, 1082, 1083, 1084, 1085, 1086, 1087, 1088,
    1089, 1090, 1091, 1092, 1093, 1094, 1095, 1096, 1097, 1098, 1099, 1100, 1101, 1102, 1103, 1104,
    1105, 1106, 1107, 1108, 1109, 1110, 1111, 1112, 1113, 1114, 1115, 1116, 1117, 1118, 1119, 1120,
    1121, 1122, 1123, 1124, 1125, 1126, 1127, 1128, 1129, 1130, 1131, 1132, 1133, 1134, 1146, 1153,
    1154, 1155, 1156, 1157, 1158, 1180, 1181, 1184, 1185, 1187, 1188, 1207, 1208, 1209, 1211, 1212
]
        self.INSTANCES = [
            # "Instances/vestel-fabrika-verisi.txt",
            # "Instances/2xVestel-synomim.txt",
            # "Instances/lrc5.txt",
            # "Instances/lrc11.txt",
            # "Instances/lrc11-location-based.txt",
            # "Instances/lrc11-demand-based.txt",
            # "Instances/lrc11-time-window-based.txt",
            # "Instances/lrc11-servtime-based.txt",
            # "Instances/lrc15.txt",
            # "Instances/lrc15-location-based.txt",
            # "Instances/lrc15-demand-based.txt",
            # "Instances/lrc15-time-window-based.txt",
            # "Instances/lrc15-servtime-based.txt",
            "Instances/lrc19.txt",
            "Instances/lrc19-location-based.txt",
            "Instances/lrc19-time-window-based.txt",
            "Instances/lrc19-servtime-based.txt",
            "Instances/lrc19-demand-based.txt"
        ]
        self.VEHICLE_COUNT = [
            1
            ,
            2
            ,
            3
        ]
        self.TROLLEY_COUNT = [
            1,
            2,
            3
        ]
        self.TROLLEY_IMPACT_RATE = [
            20,
            30,
            40
        ]
        self.EARLINESS_TARDINESS_PENALTY = [
            10,
            20,
            30
        ]

    def parse_route_math_model(self, solutionResult_MATH_MODEL, depotNode):
        resultList = list()
        for i in range(0, len(solutionResult_MATH_MODEL)):
            worksheetMathModelListItem = list()
            vehicle_info = f"Vehicle {i} - Trolley Count: {solutionResult_MATH_MODEL[i]['trolleyCount']}"
            route_math_model = ' - '.join(solutionResult_MATH_MODEL[i]['route'])

            pairs = route_math_model.split(" - ")
            pairs = [eval(pair) for pair in pairs]
            route_map = {}

            for start, end in pairs:
                route_map[start] = end

            start_point = pairs[0][0]
            for start, end in pairs:
                if start not in route_map.values():
                    start_point = start
                    break

            route = ["D0"]
            current = start_point
            while current in route_map:
                next_point = route_map[current]
                if next_point == depotNode:
                    route.append("D0")
                else:
                    route.append(f"C{next_point}")
                current = next_point

            route_str = " - ".join(map(str, route))
            worksheetMathModelListItem.append(vehicle_info)
            worksheetMathModelListItem.append(route_str)
            resultList.append(" => ".join(worksheetMathModelListItem))

        return " ; ".join(resultList)

    def cost_detail_math_model(self, routesCostDetailResult):
        distCost = sum(item[0] for item in routesCostDetailResult)
        EAPenalty = sum(item[1] for item in routesCostDetailResult if item[1] >= 0)
        TAPenalty = sum(item[2] for item in routesCostDetailResult if item[2] >= 0)
        return distCost, EAPenalty, TAPenalty

    def cost_detail_heuristic(self, solutionResult_ALNS):
        distCost = 0
        EAPenalty = 0
        TAPenalty = 0
        for k in range(0, len(solutionResult_ALNS)):
            if len(solutionResult_ALNS[k]) > 0:
                distCost += sum(item[0] for item in solutionResult_ALNS[k]["costDetail"])
                EAPenalty += sum(item[1] for item in solutionResult_ALNS[k]["costDetail"] if item[1] > 0)
                TAPenalty += sum(item[2] for item in solutionResult_ALNS[k]["costDetail"] if item[2] > 0)

        return distCost, EAPenalty, TAPenalty

    def parse_route_alns(self, solutionResult_ALNS):
        resultList = list()
        for k in range(0, len(solutionResult_ALNS)):
            if len(solutionResult_ALNS[k]) > 0:
                worksheetALNSListItem = list()
                vehicle_info_alns = f"Vehicle {k} - Trolley Count: {solutionResult_ALNS[k]['trolleyCount']}"
                route_alns = ' - '.join(solutionResult_ALNS[k]['route'])
                worksheetALNSListItem.append(vehicle_info_alns)
                worksheetALNSListItem.append(route_alns)
                resultList.append(" => ".join(worksheetALNSListItem))
        return " ; ".join(resultList)

    def parse_detail_excel(self,
                           worksheetDetail, iterationNumber, trolleyCount, trolleyImpactRate,
                           earlinessTardinessPenalty,
                           # bestCost_MATH_MODEL, cpuTime_MATH_MODEL, routesDetailResult,
                           bestCost_ALNS, cpuTime_ALNS, solutionResult_ALNS,
                           # bestCost_GA, cpuTime_GA, solutionResult_GA,
                           instance,
                           # cpuTime_RULE_OF_THUMB, bestCost_RULE_OF_THUMB, solutionResult_RULE_OF_THUMB,
                           # gapAlnsToMilp, gapAlnsToGA
                           ):
        worksheetDetail.append(
            ["ID", 'TROLLEY_COUNT', 'TROLLEY_IMPACT_RATE', 'EARLINESS/TARDINESS PENALTY'])
        worksheetDetail.append(["TEST_ID_" + str(iterationNumber), trolleyCount, trolleyImpactRate,
                                earlinessTardinessPenalty])
        worksheetDetail.append(
            [instance, "Mathematical Formulation", "ALNS", "GA", "RULE OF THUMB", "GAP (ALNS-MILP)", "GAP (ALNS-GA)"])
        worksheetDetail.append(
            ["Result", round(0), round(bestCost_ALNS), round(0), 0,
             0, 0])
        worksheetDetail.append(["cpuTime", 0, cpuTime_ALNS, 0, 0])
        # worksheetDetail.append(["Mathematical Formulation"])
        # for rd in range(0, len(routesDetailResult)):
        #     worksheetDetail.append(["", routesDetailResult[rd]])

        worksheetDetail.append(["ALNS Algorithm"])
        for k in range(0, len(solutionResult_ALNS)):
            if len(solutionResult_ALNS[k]) > 0:
                vehicle_info_alns = f"Vehicle {k} - Trolley Count: {solutionResult_ALNS[k]['trolleyCount']}"
                worksheetDetail.append(["", vehicle_info_alns])
                for rda in range(0, len(solutionResult_ALNS[k]["routeDetail"])):
                    worksheetDetail.append(["", solutionResult_ALNS[k]["routeDetail"][rda]])

        # worksheetDetail.append(["GA Algorithm"])
        # for k in range(0, len(solutionResult_GA)):
        #     if len(solutionResult_GA[k]) > 0:
        #         vehicle_info_alns = f"Vehicle {k} - Trolley Count: {solutionResult_GA[k]['trolleyCount']}"
        #         worksheetDetail.append(["", vehicle_info_alns])
        #         for rda in range(0, len(solutionResult_GA[k]["routeDetail"])):
        #             worksheetDetail.append(["", solutionResult_GA[k]["routeDetail"][rda]])
        #
        # worksheetDetail.append(["Rule Of Thumb Algorithm"])
        # for kr in range(0, len(solutionResult_RULE_OF_THUMB)):
        #     if len(solutionResult_RULE_OF_THUMB[kr]) > 0:
        #         vehicle_info_alns = f"Vehicle {kr} - Trolley Count: {solutionResult_RULE_OF_THUMB[kr]['trolleyCount']}"
        #         worksheetDetail.append(["", vehicle_info_alns])
        #         for rdaT in range(0, len(solutionResult_RULE_OF_THUMB[kr]["routeDetail"])):
        #             worksheetDetail.append(["", solutionResult_RULE_OF_THUMB[kr]["routeDetail"][rdaT]])

        worksheetDetail.append([])

    def parse_roth_detail_excel(self, worksheetRot, file_name, vehicleCount, iterationNumber, trolleyCount,
                                trolleyImpactRate, earlinessTardinessPenalty,
                                minResult, rule_of_thumb_dist_cost, rule_of_thumb_EACost, rule_of_thumb_TAcost,
                                rule_of_thumb_routes, bestCost_RULE_OF_THUMB, bestDistance_TWO,
                                bestDistance_THREE, bestDistance_FOUR, milpresult, alnsresult
                                ):
        worksheetRot.append(
            [
                iterationNumber, file_name, vehicleCount, trolleyCount, trolleyImpactRate,
                earlinessTardinessPenalty,
                minResult[2], round(minResult[1]),
                round(rule_of_thumb_dist_cost),
                round(rule_of_thumb_EACost), round(rule_of_thumb_TAcost), rule_of_thumb_routes,
                minResult[0],
                round(bestCost_RULE_OF_THUMB),
                round(bestDistance_TWO),
                round(bestDistance_THREE),
                round(bestDistance_FOUR),
                0,
                0
            ])

    def execute(self):
        start_time = time.time()
        print("Scenario Started")
        iterationNumber = 810
        iterationCount = 3500
        capacityOfTrolley = 60
        timeLimit = 60000 * 20
        file_name = "result-all-alns-with-mlp-3"
        workbook = Workbook()
        workbook.remove(workbook.active)  # Remove default sheet
        worksheet = workbook.create_sheet(title=f'Test Result')
        worksheet.append(
            [
                "", "TEST INSTANCES", "", "", "", "",
                # "MATHEMATICAL FORMULATION", "", "", "", "", "", "",
                "ALNS", "", "", "", "", "",
                # "Genetic Algorithm", "", "", "", "", "",
                # "Min Rule Of Thumb", "", "", "", "", "", "",
                # "GAP (ALNS-MILP)",
                # "GAP (ALNS-GA)",
            ])
        worksheet.append(
            [
                "TEST ID", "FILE NAME", "VEHICLE COUNT", 'TROLLEY COUNT', 'TROLLEY IMPACT TIME',
                "EARLINESS/TARDINESS PENALTY",
                # "Math Form Is Optimal OR Feasible", "CPU TIME", "RESULT", "Math Model Distance Cost",
                # "Math Model EA Cost", "Math Model TA Cost", "ROUTES",
                "CPU TIME", "RESULT",
                "ALNS Distance Cost", "ALNS EA Cost", "ALNS TA Cost", "ROUTES",
                # "CPU TIME", "RESULT",
                # "GA Distance Cost", "GA EA Cost", "GA TA Cost", "ROUTES",
                # "CPU TIME", "RESULT",
                # "ROTH Distance Cost", "ROTH EA Cost", "ROTH TA Cost", "ROUTES", "MIN ALGO NAME",
                # "",
                # ""
            ])
        infeasibleData = list()
        nonOptimalSolution = list()
        workbookDetail = Workbook()
        workbookDetail.remove(workbookDetail.active)  # Remove default sheet

        workbookROTDetail = Workbook()
        workbookROTDetail.remove(workbookROTDetail.active)  # Remove default sheet
        worksheetRot = workbookROTDetail.create_sheet(title=f'Test Result')
        # worksheetRot.append(
        #     [
        #         "", "TEST INSTANCES", "", "", "", "",
        #         "Min Rule Of Thumb", "", "", "", "", "",
        #         "", "", "", "", "", "", ""
        #     ])
        # worksheetRot.append(
        #     [
        #         "TEST ID", "FILE NAME", "VEHICLE COUNT", 'TROLLEY COUNT', 'TROLLEY IMPACT TIME',
        #         "EARLINESS/TARDINESS PENALTY",
        #         "CPU TIME", "RESULT",
        #         "ROTH Distance Cost", "ROTH EA Cost", "ROTH TA Cost", "ROUTES", "MIN ALGO NAME",
        #         "SORTED BY COST", "GREEDY", "SORTED BY DIST", "SORTED BY DUE DATE", "MILP Result", "ALNS result"
        #     ])


        for ins in range(0, len(self.INSTANCES)):
            instance = self.INSTANCES[ins]
            fileName = instance.split('/')[-1].replace('.txt', '')
            for vc in range(0, len(self.VEHICLE_COUNT)):
                vehicleCount = self.VEHICLE_COUNT[vc]
                worksheetDetail = workbookDetail.create_sheet(title=f'{vehicleCount} Vehicle')
                for tc in range(0, len(self.TROLLEY_COUNT)):
                    trolleyCount = self.TROLLEY_COUNT[tc]
                    for tir in range(0, len(self.TROLLEY_IMPACT_RATE)):
                        trolleyImpactRate = self.TROLLEY_IMPACT_RATE[tir]
                        for etp in range(0, len(self.EARLINESS_TARDINESS_PENALTY)):
                            iterationNumber += 1
                            earlinessTardinessPenalty = self.EARLINESS_TARDINESS_PENALTY[etp]
                            print(
                                "===================================================================================================================================")
                            print("Iteration Number : " + str(iterationNumber))
                            print(
                                f"INSTANCE: {instance}, VEHICLE_COUNT : {vehicleCount}, TROLLEY_COUNT: {trolleyCount}"
                                f", TROLLEY_IMPACT_RATE: {trolleyImpactRate}, EARLINESS_TARDINESS_PENALTY: {earlinessTardinessPenalty}, "
                                f"POPULATION SIZE: {40}, SELECTION SIZE: {8}")

                            # print("==================== MILP START =====================")

                            if iterationNumber not in self.initeration:
                                continue

                            # algorithm = Algorithm(instance, vehicleCount, capacityOfTrolley, trolleyCount,
                            #                       trolleyImpactRate, earlinessTardinessPenalty,
                            #                       earlinessTardinessPenalty, timeLimit)
                            # optimalOrFeasible, bestCost_MATH_MODEL, cpuTime_MATH_MODEL, solutionResult_MATH_MODEL, depotNode, routesDetailResult, routesCostDetailResult = algorithm.execute()
                            #
                            # if optimalOrFeasible == "NON-OPTIMAL":
                            #     nonOptimalSolution.append(
                            #         ["TEST_ID: " + str(iterationNumber),
                            #          "INSTANCE: " + str(
                            #              instance) + ", VEHICLE_COUNT: " + str(
                            #              vehicleCount) + ", TROLLEY_COUNT: " + str(
                            #              trolleyCount) + ", TROLLEY_IMPACT_RATE: " + str(
                            #              trolleyImpactRate) + ", EARLINESS_TARDINESS_PENALTY: " + str(
                            #              earlinessTardinessPenalty)
                            #          ])

                            # print("==================== ALNS START =====================")
                            heuristic = Heuristic(instance, vehicleCount, capacityOfTrolley, trolleyCount,
                                                  trolleyImpactRate, earlinessTardinessPenalty,
                                                  earlinessTardinessPenalty, iterationCount, 100, file_name, fileName)
                            bestCost_ALNS, cpuTime_ALNS, solutionResult_ALNS = heuristic.execute()
                            # print("==================== GA START =======================")
                            # genetic = Genetic(instance, vehicleCount, capacityOfTrolley, trolleyCount,
                            #                   trolleyImpactRate, earlinessTardinessPenalty,
                            #                   earlinessTardinessPenalty, iterationCount, 10, 2)
                            # bestCost_GA, cpuTime_GA, solutionResult_GA = genetic.execute()

                            # print("==================== ROTH START =====================")
                            # ruleOfThumb = RuleOfThumb(instance, vehicleCount, capacityOfTrolley, trolleyCount,
                            #                           trolleyImpactRate, earlinessTardinessPenalty,
                            #                           earlinessTardinessPenalty, 0, 100)
                            # bestCost_RULE_OF_THUMB, bestDistance_TWO, bestDistance_THREE, bestDistance_FOUR, minResult = ruleOfThumb.execute()

                            # gapAlnsToMilp = ((round(bestCost_MATH_MODEL) - round(bestCost_ALNS)) / round(
                            #     bestCost_ALNS)) * 100
                            # if gapAlnsToMilp > 1 or gapAlnsToMilp < 0:
                            #     infeasibleData.append(
                            #         ["TEST_ID: " + str(iterationNumber), "INSTANCE: " + str(
                            #             instance) + ", VEHICLE_COUNT: " + str(
                            #             vehicleCount) + ", TROLLEY_COUNT: " + str(
                            #             trolleyCount) + ", TROLLEY_IMPACT_RATE: " + str(
                            #             trolleyImpactRate) + ", EARLINESS_TARDINESS_PENALTY: " + str(
                            #             earlinessTardinessPenalty), gapAlnsToMilp, "ALNS TO MILP"])
                            #
                            # gapAlnsToGA = ((round(bestCost_GA) - round(bestCost_ALNS)) / round(
                            #     bestCost_ALNS)) * 100
                            # if gapAlnsToGA > 1 or gapAlnsToGA < 0:
                            #     infeasibleData.append(
                            #         ["TEST_ID: " + str(iterationNumber), "INSTANCE: " + str(
                            #             instance) + ", VEHICLE_COUNT: " + str(
                            #             vehicleCount) + ", TROLLEY_COUNT: " + str(
                            #             trolleyCount) + ", TROLLEY_IMPACT_RATE: " + str(
                            #             trolleyImpactRate) + ", EARLINESS_TARDINESS_PENALTY: " + str(
                            #             earlinessTardinessPenalty), gapAlnsToGA, "ALNS TO GA"])

                            # math_model_routes = self.parse_route_math_model(solutionResult_MATH_MODEL, depotNode)
                            heuristic_routes = self.parse_route_alns(solutionResult_ALNS)
                            # ga_routes = self.parse_route_alns(solutionResult_GA)
                            # rule_of_thumb_routes = self.parse_route_alns(minResult[3])

                            # mathModel_dist_cost, mathModel_EACost, mathModel_TAcost = self.cost_detail_math_model(
                            #     routesCostDetailResult)
                            heuristic_dist_cost, heuristic_EACost, heuristic_TAcost = self.cost_detail_heuristic(
                                solutionResult_ALNS)
                            # ga_dist_cost, ga_EACost, ga_TAcost = self.cost_detail_heuristic(
                            #     solutionResult_GA)
                            # rule_of_thumb_dist_cost, rule_of_thumb_EACost, rule_of_thumb_TAcost = self.cost_detail_heuristic(
                            #     minResult[3])
                            worksheet.append(
                                [
                                    iterationNumber, fileName, vehicleCount, trolleyCount, trolleyImpactRate,
                                    earlinessTardinessPenalty,
                                    # optimalOrFeasible,
                                    # cpuTime_MATH_MODEL, round(bestCost_MATH_MODEL), round(mathModel_dist_cost),
                                    # round(mathModel_EACost), round(mathModel_TAcost),
                                    # math_model_routes,
                                    cpuTime_ALNS, round(bestCost_ALNS), round(heuristic_dist_cost),
                                    round(heuristic_EACost), round(heuristic_TAcost), heuristic_routes,
                                    # cpuTime_GA, round(bestCost_GA), round(ga_dist_cost),
                                    # round(ga_EACost), round(ga_TAcost), ga_routes,
                                    # minResult[2], round(minResult[1]),
                                    # round(rule_of_thumb_dist_cost),
                                    # round(rule_of_thumb_EACost), round(rule_of_thumb_TAcost), rule_of_thumb_routes,
                                    # minResult[0],
                                    # gapAlnsToMilp,
                                    # gapAlnsToGA

                                ])

                            self.parse_detail_excel(worksheetDetail, iterationNumber, trolleyCount,
                                                    trolleyImpactRate, earlinessTardinessPenalty,
                                                    # bestCost_MATH_MODEL,  cpuTime_MATH_MODEL, routesDetailResult,
                                                    bestCost_ALNS, cpuTime_ALNS, solutionResult_ALNS,
                                                    # bestCost_GA, cpuTime_GA, solutionResult_GA,
                                                    instance,
                                                    # minResult[2], minResult[1],
                                                    # minResult[3],
                                                    # gapAlnsToMilp,
                                                    # gapAlnsToGA
                                                    )
                            # self.parse_roth_detail_excel(worksheetRot, file_name, vehicleCount, iterationNumber,
                            #                              trolleyCount,
                            #                              trolleyImpactRate, earlinessTardinessPenalty,
                            #                              minResult,
                            #                              round(rule_of_thumb_dist_cost),
                            #                              round(rule_of_thumb_EACost), round(rule_of_thumb_TAcost),
                            #                              rule_of_thumb_routes,
                            #                              round(bestCost_RULE_OF_THUMB),
                            #                              round(bestDistance_TWO),
                            #                              round(bestDistance_THREE),
                            #                              round(bestDistance_FOUR), round(0),
                            #                              round(0))
            workbookDetail.save(f'excels/details/{fileName}-details.xlsx')
            workbookROTDetail.save(f'excels/rot-detail/{file_name}-roth-details.xlsx')

        # if len(infeasibleData) > 0:
        #     workbookInFeasible = Workbook()
        #     workbookInFeasible.remove(workbookInFeasible.active)
        #     worksheetInFeasible = workbookInFeasible.create_sheet(title="Gap Data")
        #     worksheetInFeasible.append(["TEST_ID", "TEST INSTANCE", "GAP"])
        #     for inf in range(0, len(infeasibleData)):
        #         worksheetInFeasible.append(infeasibleData[inf])
        #     workbookInFeasible.save(f'excels/gap/{file_name}-gap.xlsx')
        #
        #
        # if len(nonOptimalSolution) > 0:
        #     workbookNonOptimal = Workbook()
        #     workbookNonOptimal.remove(workbookNonOptimal.active)
        #     worksheetNonOptimal = workbookNonOptimal.create_sheet(title="Non-optimal Data")
        #     worksheetNonOptimal.append(["TEST_ID", "TEST INSTANCE"])
        #     for f in range(0, len(nonOptimalSolution)):
        #         worksheetNonOptimal.append(nonOptimalSolution[f])
        #     workbookNonOptimal.save(f'excels/non-optimal/{file_name}-non-optimal.xlsx')

        workbook.save(f'excels/{file_name}.xlsx')
        print("Scenario finished")
        endtime = time.time()
        cpuTime = round(endtime - start_time, 1)

        print("cpuTime: " + str(cpuTime) + " seconds")


scenario = ScenarioAnalysis()
scenario.execute()
